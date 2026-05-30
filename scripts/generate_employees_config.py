#!/usr/bin/env python3
"""Generate config/employees.php: urutan nomor lama (1-40) + data NIP dari Excel."""

import openpyxl
from pathlib import Path

EXCEL = Path(__file__).resolve().parents[2] / "NOMOR INDUK KARYAWAN TOEDJOE SINAR GROUP (2).xlsx"

EXCLUDE_NAMES = {"ELGI SAMSUDDIN"}

OLD_ORDER = [
    "SUGIANTO",
    "MUHAMMAD NOR DIANSYAH",
    "SRI WAHYUNI",
    "FELIX OCTAVIAN MAHENDRA",
    "SUTIKNO",
    "DEKI SAPUTRA",
    "ISMAIL",
    "MUHAJIR",
    "MOHAMMAD NADZORI",
    "SANDRA WIDIYA",
    "MUHAMMAD IHSANUL AKMAL",
    "AGUNG DWI RAHMADI",
    "ANJASMARA RAMADHAN",
    "FANI NOR HAMDANI",
    "IKE TRIWAHYUNI",
    "AHMAD RIDHANI",
    "MUHAMMAD FERDIAN",
    "ZUASA RAZZAQ GHANI",
    "ARYA GYMNAS FATAHILAH",
    "RISNUR RAMADHAN",
    "MAULANA RUMI DAENG RINGGI",
    "MUHAMMAD FAUZAN",
    "NUR RINO AIDUL FADLI",
    "RUSMONO HADI",
    "MUHAMMAD YUSRIL ARIMULIA",
    "MUHAMMAD ZIQI FARHAN",
    "RISKI NUR HALIS",
    "AHMAT CAHYADI",
    "MUHAMMAD RAWINDRA FADILAH",
    "MUHAMMAD KAMAL",
    "MUHAMMAD ILHAM NAZHIF",
    "AYU NANDA WULANDARI",
    "ATHASYAHRI SYAWAL FAHREZY",
    "DAVI JUNIAR",
    "MUHAMMAD ZIKRI FARHAN",
    "MUHAMMAD NAUFAL PRIMANANDA",
    "MUHAMMAD ZAKI MUBARAK",
    "MUHAMMAD RAFI'I",
    "IMROATUS SHOLIKHA",
    "EFIN QOID AMALLUDIN",
]

FALLBACK = {
    "SUGIANTO": ("sugynuansafx@gmail.com", "Kepala Maintenance Operator", "Samarinda", "2015-08-01"),
    "MUHAMMAD NOR DIANSYAH": ("Nurdinzds@gmail.com", "Kepala Operator Indoor", "Samarinda", "2015-08-01"),
    "SRI WAHYUNI": ("uyung109@gmail.com", "Kepala Desainer", "Samarinda", "2015-08-01"),
    "FELIX OCTAVIAN MAHENDRA": ("mahendrafelix@gmail.com", "Kepala Admin & Personalia", "Samarinda", "2017-08-01"),
    "SUTIKNO": ("sutiknokosong7@gmail.com", "Kepala Finishing", "Samarinda", "2017-08-01"),
    "DEKI SAPUTRA": ("pandabakpao10@gmail.com", "Kepala TEXTILE/DTF", "Samarinda", "2017-10-02"),
    "ISMAIL": ("mail87.ismail@gmail.com", "Asisten Operator Indoor", "Samarinda", "2018-05-01"),
    "MUHAJIR": ("ajirgalaxy@gmail.com", "Kepala Laser/UV/DTF", "Samarinda", "2019-05-22"),
    "MOHAMMAD NADZORI": ("m.nadzory@gmail.com", "Asisten Operator Indoor", "Samarinda", "2019-11-22"),
    "SANDRA WIDIYA": ("sandrawdy40@gmail.com", "Design & Kepala Kasir", "Samarinda", "2020-08-01"),
    "MUHAMMAD IHSANUL AKMAL": ("muhammadihsanulakmal.18@gmail.com", "Kepala Operator Digital A3", "Samarinda", "2020-08-01"),
    "AGUNG DWI RAHMADI": ("agungdr463@gmail.com", "Asisten Operator Indoor", "Samarinda", "2020-09-01"),
    "ANJASMARA RAMADHAN": ("anjasoedinson@gmail.com", "Asisten Digital A3", "Samarinda", "2021-01-01"),
    "FANI NOR HAMDANI": ("faninh41@gmail.com", "Asisten Desainer", "Samarinda", "2021-11-28"),
    "IKE TRIWAHYUNI": ("triwahyuniike50@gmail.com", "Asisten Desainer & Kasir", "Samarinda", "2022-07-01"),
    "AHMAD RIDHANI": ("ahmadridhani87@gmail.com", "Asisten Desainer", "Samarinda", "2022-07-01"),
    "MUHAMMAD FERDIAN": ("muhammadferdian0451@gmail.com", "Asisten Desainer", "Samarinda", "2022-08-01"),
    "ZUASA RAZZAQ GHANI": ("zuasarazak1@gmail.com", "Asisten Desainer", "Samarinda", "2022-09-01"),
    "ARYA GYMNAS FATAHILAH": ("aryagymnasf25@gmail.com", "Asisten Textile/DTF", "Samarinda", "2022-10-01"),
    "RISNUR RAMADHAN": ("ramadhanrisnur@gmail.com", "Asisten Desainer", "Samarinda", "2023-05-20"),
    "MAULANA RUMI DAENG RINGGI": ("itsmeschweigen@gmail.com", "Asisten Operator Outdoor", "Samarinda", "2023-05-15"),
    "MUHAMMAD FAUZAN": ("m.f4uzan12477@gmail.com", "Asisten Operator Outdoor", "Samarinda", "2023-06-06"),
    "NUR RINO AIDUL FADLI": ("rinofadlipratama04@gmail.com", "Kepala Konten & Marketing", "Samarinda", "2023-06-26"),
    "RUSMONO HADI": ("Rusmonohady@gmail.com", "Asisten Operator Outdoor", "Samarinda", "2023-08-09"),
    "MUHAMMAD YUSRIL ARIMULIA": ("sayangyusril22@gmail.com", "Kepala Pengambilan", "Samarinda", "2023-12-15"),
    "MUHAMMAD ZIQI FARHAN": ("ziqigustinaldi@gmail.com", "Asisten Operator Outdoor", "Samarinda", "2024-01-29"),
    "RISKI NUR HALIS": ("riskinurhalis11@gmail.com", "Asisten Laser/UV/DTF", "Samarinda", "2024-02-02"),
    "AHMAT CAHYADI": ("cahyadiahmat99@gmail.com", "Asisten Pengambilan", "Samarinda", "2024-05-01"),
    "MUHAMMAD RAWINDRA FADILAH": ("Fadilfira31@gmail.com", "Asisten Digital A3", "Samarinda", "2024-05-20"),
    "MUHAMMAD KAMAL": ("raeelyy21@gmail.com", "Asisten Merchandise", "Samarinda", "2024-06-03"),
    "MUHAMMAD ILHAM NAZHIF": ("m.ilhamnazhif23@gmail.com", "Asisten Operator Outdoor", "Samarinda", "2024-09-16"),
    "AYU NANDA WULANDARI": ("ayunandawul15@gmail.com", "Desainer Admin Online", "Samarinda", "2024-12-23"),
    "ATHASYAHRI SYAWAL FAHREZY": ("athafarez2412@gmail.com", "Website", "Samarinda", "2025-01-01"),
    "DAVI JUNIAR": ("juniardavi5@gmail.com", "Asisten Digital A3", "Samarinda", "2025-02-10"),
    "MUHAMMAD ZIKRI FARHAN": ("zikrifarhan05@gmail.com", "Asisten Indoor Cutting", "Samarinda", "2025-05-03"),
    "MUHAMMAD NAUFAL PRIMANANDA": ("az5733356@gmail.com", "Asisten Pengambilan", "Samarinda", "2025-05-12"),
    "MUHAMMAD ZAKI MUBARAK": ("zm0700910@gmail.com", "Asisten Pengambilan", "Samarinda", "2025-05-13"),
    "MUHAMMAD RAFI'I": ("rfyxyz@gmail.com", "Asisten Desainer", "Samarinda", "2025-05-04"),
    "IMROATUS SHOLIKHA": ("imroatussholikha8597@gmail.com", "Asisten Desainer", "Samarinda", "2025-10-27"),
    "EFIN QOID AMALLUDIN": ("efinqoidamalludin@gmail.com", "Asisten Desainer", "Samarinda", "2025-11-03"),
}


def php_str(value: str) -> str:
    return "'" + value.replace("\\", "\\\\").replace("'", "\\'") + "'"


def load_excel() -> dict[str, dict]:
    wb = openpyxl.load_workbook(EXCEL, data_only=True)
    ws = wb.active
    rows: dict[str, dict] = {}

    for r in range(3, ws.max_row + 1):
        name = str(ws.cell(r, 3).value or "").strip().upper()
        if not name or name in EXCLUDE_NAMES:
            continue

        birth_year = int(ws.cell(r, 4).value)
        birth_month = str(ws.cell(r, 5).value).zfill(2)
        birth_day = str(ws.cell(r, 6).value).zfill(2)
        address = ws.cell(r, 11).value

        rows[name] = {
            "tgl_lahir": f"{birth_year}-{birth_month}-{birth_day}",
            "jenis_kelamin": str(ws.cell(r, 9).value).strip(),
            "alamat": (address or "Samarinda").strip(),
        }

    return rows


def build_nip(nomor: int, tgl_masuk: str, tgl_lahir: str) -> str:
    hy, _, _ = tgl_masuk.split("-")
    by, bm, bd = tgl_lahir.split("-")
    return f"TSG-15-{hy}-{bd}-{bm}-{by}-{nomor:02d}"


def main() -> None:
    excel = load_excel()
    lines = []

    for nomor, name in enumerate(OLD_ORDER, start=1):
        fb = FALLBACK[name]
        ex = excel.get(name, {})

        email, jabatan, alamat_fb, tgl_masuk_fb = fb
        tgl_lahir = ex.get("tgl_lahir")
        tgl_masuk = tgl_masuk_fb
        jenis_kelamin = ex.get("jenis_kelamin", "LAKI-LAKI")
        alamat = ex.get("alamat", alamat_fb)

        if not tgl_lahir:
            raise SystemExit(f"Data lahir tidak ditemukan di Excel untuk: {name}")

        nip = build_nip(nomor, tgl_masuk, tgl_lahir)

        lines.append(
            f"        {nomor} => ['nip' => {php_str(nip)}, 'name' => {php_str(name)}, "
            f"'email' => {php_str(email)}, 'jabatan' => {php_str(jabatan)}, "
            f"'alamat' => {php_str(alamat)}, 'tgl_lahir' => {php_str(tgl_lahir)}, "
            f"'tgl_masuk' => {php_str(tgl_masuk)}, 'jenis_kelamin' => {php_str(jenis_kelamin)}],"
        )

    print("\n".join(lines))


if __name__ == "__main__":
    main()
